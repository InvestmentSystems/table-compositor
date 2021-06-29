from functools import lru_cache

from openpyxl.styles import (
    Alignment,
    Border,
    Color,
    Font,
    PatternFill,
    Side,
    fills,
)

SENTINEL = object()


class OpenPyxlStyleHelper:

    # good reference:
    # https://www.ozgrid.com/Excel/excel-custom-number-formats.htm
    DOLLAR_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    CENTS_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # without dollar sign, but has same rendering format for currency
    GENERAL_CURRENCY_FORMAT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    PERCENT_FORMAT = "0.00%"

    class CustomBorders:
        thin_white = Side(border_style="thin", color="FFFFFF")
        thick_white = Side(border_style="thick", color="FFFFFF")

        thin_black = Side(border_style="thin", color="000000")
        thick_black = Side(border_style="thick", color="000000")

        border = Border(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
        )
        left_border = Border(
            left=thick_white, right=thin_white, top=thin_white, bottom=thin_white
        )
        right_border = Border(
            left=thin_white, right=thick_white, top=thin_white, bottom=thin_white
        )

        top_right_border = Border(
            left=thin_white, right=thin_black, top=thin_black, bottom=thin_white
        )
        top_left_border = Border(
            left=thin_white, right=thin_black, top=thin_black, bottom=thin_white
        )

        top_border = Border(
            left=thin_white, right=thin_white, top=thin_black, bottom=thin_white
        )

        bottom_right_border = Border(
            left=thin_white, right=thin_black, top=thin_white, bottom=thin_black
        )
        bottom_left_border = Border(
            left=thin_black, right=thin_white, top=thin_white, bottom=thin_black
        )
        bottom_border = Border(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_black
        )

        thin_border = Border(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
        )

        thin_black_border = Border(
            left=thin_black, right=thin_black, top=thin_black, bottom=thin_black
        )
        thin_white_border = Border(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
        )
        thick_side_border = Border(
            left=thick_white, right=thick_white, top=thin_white, bottom=thin_white
        )

    @staticmethod
    @lru_cache(maxsize=None)
    def default_header_style(
        *,
        alignment="center",
        font=Font(bold=True, color="FFFFFF"),
        bgColor="4F81BD",
        border=SENTINEL,
    ):

        """
        Provides styles for default headers for OpenPyxl engine

        Args:
            alignment: 'center', 'left', 'right' used for horizontal alignment
            font: an openpyxl.Font instance
            bgColor: hex color that will be used as background color in the fill pattern
            border: an openpyxl.Border instance, defaults to thin white border

        Returns:
            A dict of key-values pairs, where each key is a attr of the `cell` object in openyxl and value is valid value of that attr.
        """

        if border is SENTINEL:
            border = OpenPyxlStyleHelper.CustomBorders.thin_white_border

        return dict(
            alignment=Alignment(horizontal=alignment),
            font=font,
            fill=PatternFill(patternType="solid", fgColor=bgColor),
            border=border,
        )

    @staticmethod
    @lru_cache(maxsize=None)
    def get_style(number_format="General", bg_color=None, border=None, font=None):

        """
        Helper method to return a openpyxl Style

        Args:
            number_format: an xlsx compatibale number format string
            bg_color: hex color that will be used as background color in the fill pattern
            border: an openpyxl.Border instance, defaults to thin white border

        Returns:
            A dict of key-values pairs, where each key is a attr of the `cell` object in openyxl and value is valid value of that attr.
        """
        pattern_fill = PatternFill()
        if bg_color:
            pattern_fill = PatternFill(
                fgColor=Color(bg_color), patternType=fills.FILL_SOLID
            )

        border = border or Border()
        result = dict(fill=pattern_fill, border=border, number_format=number_format)
        if font is not None:
            result["font"] = font
        return result


class XLSXWriterDefaults:
    """
    Class provides defaults callback funcs that can be used
    while calling the build_presentation_model.
    """

    @staticmethod
    def data_value_func(df):
        """
        Default value that can be used as callback for data_value_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            A function that takes idx, col as arguments and returns the df.loc[idx, col] value
        """

        def _value_func(r, c):
            return df.loc[r, c]

        return _value_func

    @staticmethod
    def data_style_func(df):
        """
        Default value that can be used as callback for data_style_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            a function table takes idx, col as arguments and returns a openpyxl compatible style dictionary
        """

        def _style_func(r, c):
            return OpenPyxlStyleHelper.get_style()

        return _style_func

    @staticmethod
    def header_value_func(df):
        """
        Default value that can be used as callback for data_header_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            a function table takes `node` as arguments and returns node.value
        """

        def _value_func(node):
            return node.value

        return _value_func

    @staticmethod
    def header_style_func(df):
        """
        Default value that can be used as callback for data_style_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            a function table takes `node` as arguments and returns a openpyxl compatible style dictionary
        """

        def _style_func(node):
            return OpenPyxlStyleHelper.default_header_style()

        return _style_func

    @staticmethod
    def index_value_func(df):
        """
        Default value that can be used as callback for index_header_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            a function table takes `node` as arguments and returns node.value
        """

        def _value_func(node):
            return node.value

        return _value_func

    @staticmethod
    def index_style_func(df):
        """
        Default value that can be used as callback for index_style_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            a function table takes `node` as arguments and returns a openpyxl compatible style dictionary
        """

        def _style_func(node):
            return OpenPyxlStyleHelper.get_style()

        return _style_func

    @staticmethod
    def index_name_style_func(df):
        """
        Default value that can be used as callback for index_name_style_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            a function table takes index.name as arguments and returns a openpyxl compatible style dictionary"""

        def _style_func(index_name):
            return OpenPyxlStyleHelper.default_header_style()

        return _style_func

    @staticmethod
    def index_name_value_func(df):
        """
        Default value that can be used as callback for index_name_value_func

        Args:
            df: the dataframe that will be used to build the presentation model

        Returns:
            a function table takes index.name as arguments and returns index.name if not None, else ''
        """

        def _value_func(index_name):
            return index_name or ""

        return _value_func


class XlsxWriterStyleHelper:
    """
    Class provides style objects for XlsxWriter library uses to render xlsx files
    """

    # good reference:
    # https://www.ozgrid.com/Excel/excel-custom-number-formats.htm
    DOLLAR_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    CENTS_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    # without dollar sign, but has same rendering format for currency
    GENERAL_CURRENCY_FORMAT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    PERCENT_FORMAT = "0.00%"

    class CustomBorders:

        thin_white = dict(border_style=1, color="#FFFFFF")
        thick_white = dict(border_style=2, color="#FFFFFF")

        thin_black = dict(border_style=1, color="#000000")

        border = dict(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
        )
        left_border = dict(
            left=thin_black, right=thin_white, top=thin_white, bottom=thin_white
        )
        right_border = dict(
            left=thin_white, right=thick_white, top=thin_white, bottom=thin_white
        )

        top_right_border = dict(
            left=thin_white, right=thin_black, top=thin_black, bottom=thin_white
        )
        top_left_border = dict(
            left=thin_white, right=thin_black, top=thin_black, bottom=thin_white
        )

        top_border = dict(
            left=thin_white, right=thin_white, top=thin_black, bottom=thin_white
        )

        bottom_right_border = dict(
            left=thin_white, right=thin_black, top=thin_white, bottom=thin_black
        )
        bottom_left_border = dict(
            left=thin_black, right=thin_white, top=thin_white, bottom=thin_black
        )
        bottom_border = dict(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_black
        )

        thin_border = dict(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
        )

        thin_black_border = dict(
            left=thin_black, right=thin_black, top=thin_black, bottom=thin_black
        )

        thin_white_border = dict(
            left=thin_white, right=thin_white, top=thin_white, bottom=thin_white
        )

    @staticmethod
    def default_header_style(
        *,
        number_format="General",
        alignment="center",
        font=None,
        bgColor="#4F81BD",
        border=SENTINEL,
    ):

        """
        Provides styles for default headers for XlsxWriter engine

        Args:
            alignment: 'center', 'left', 'right' used for horizontal alignment
            font: an openpyxl.Font instance
            bgColor: hex color that will be used as background color in the fill pattern
            border: an openpyxl.Border instance, defaults to thin white border

        Returns:
            A dict of key-values pairs, where each key is a attr of the `cell` object in openyxl and value is valid value of that attr.
        """

        if not font:
            font = dict(bold=True, color="#FFFFFF")

        if border is SENTINEL:
            border = XlsxWriterStyleHelper.CustomBorders.thin_white_border

        return dict(
            num_format=number_format,
            align=alignment,
            valign="vcenter",
            bold=font["bold"],
            font_color=font["color"],
            pattern=1,  # 1 is solid in XlsxWriter
            bg_color=bgColor,
            top=border["top"]["border_style"],
            left=border["left"]["border_style"],
            bottom=border["bottom"]["border_style"],
            right=border["bottom"]["border_style"],
            top_color=border["top"]["color"],
            left_color=border["left"]["color"],
            bottom_color=border["bottom"]["color"],
            right_color=border["right"]["color"],
        )

    @staticmethod
    def get_style(number_format="General", bg_color=None, border=SENTINEL):

        """
        Helper method to return Style dictionary for XlsxWriter engine

        Args:
            number_format: an xlsx compatibale number format string
            bg_color: hex color that will be used as background color in the fill pattern
            border: an openpyxl.Border instance, defaults to thin white border

        Returns:
            A dict of key-values pairs, where each key/value is compatible with the `Format` object in XlsxWriter library.
        """
        styles = dict(num_format=number_format)

        if bg_color:
            styles["pattern"] = 1
            styles["bg_color"] = bg_color

        if border is SENTINEL:
            border = XlsxWriterStyleHelper.CustomBorders.thin_white_border

        border_attrs = dict(
            top=border["top"]["border_style"],
            left=border["left"]["border_style"],
            bottom=border["bottom"]["border_style"],
            right=border["right"]["border_style"],
            top_color=border["top"]["color"],
            left_color=border["left"]["color"],
            bottom_color=border["bottom"]["color"],
            right_color=border["right"]["color"],
        )
        styles.update(border_attrs)

        return styles

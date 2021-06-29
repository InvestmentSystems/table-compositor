import functools
import warnings
from itertools import chain

import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from table_compositor.grid import GridLayoutManager
from table_compositor.util import df_type_to_str

_DEFAULT_COLUMN_WIDTH = 20


class _XLSXCompositor:
    """
    Base class that can be customized for different xlsx writer engines
    """

    @staticmethod
    @functools.lru_cache(maxsize=None)
    def _get_column_letter(col):
        return get_column_letter(col)

    @classmethod
    def _build_row_col_dict(cls, layout, orientation, h_shift_by, v_shift_by):
        row_col_dict = GridLayoutManager.get_row_col_dict(
            layout,
            orientation=orientation,
            h_shift_by=h_shift_by,
            v_shift_by=v_shift_by,
        )
        return row_col_dict

    @classmethod
    def to_xlsx_worksheet(self, *args, **kwargs):
        raise NotImplementedError(
            "Derived classes need to implement this function depending on the xlsx writer engine"
        )

    @classmethod
    def to_xlsx(cls, *args, **kwargs):
        raise NotImplementedError(
            "Derived classes need to implement this function depending on the xlsx writer engine"
        )


class OpenPyxlCompositor(_XLSXCompositor):
    """
    Class provides functions to render xlsx files using the `openpyxl` library.
    """

    RANGE_STRING = "{}{}:{}{}"

    @classmethod
    @functools.lru_cache(maxsize=None)
    def _get_range_string(cls, offsets):
        r1, c1, r2, c2 = offsets

        range_string = cls.RANGE_STRING.format(
            cls._get_column_letter(c1), r1, cls._get_column_letter(c2), r2
        )
        return range_string

    @classmethod
    def _to_xlsx_worksheet(cls, row_col_dict, ws, column_width, post_process_ws_func):
        """
        creates the excel sheet using openpyxl
        """
        for offsets, (value, style, _) in row_col_dict.items():
            offsets = tuple(i + 1 for i in offsets)  # bump needed for openpyxl
            cell = ws.cell(
                row=offsets[0], column=offsets[1], value=df_type_to_str(value)
            )
            if offsets[0] != offsets[2] or offsets[1] != offsets[3]:
                ws.merge_cells(
                    start_row=offsets[0],
                    start_column=offsets[1],
                    end_row=offsets[2],
                    end_column=offsets[3],
                )

            rows = ws[cls._get_range_string(offsets)]
            # we do this since some formatting (fill, border) for merged cells still need to be set for each individual cell. we assume setting the same formatting (eg. border) for each cell in merged range will mask the values such that it feels like the merged cell is being formatted as one cell. for example the left border  is derived from leftmost cell and right border is derived from right most cell, even though same border attributes are set for each cell in the merged range. for any in-between cells the border formatting will be voided.
            for cell in chain(*rows):
                for attr, style_value in style.user_style.items():
                    try:
                        setattr(cell, attr, style_value)
                    except AttributeError:
                        # we do not set the attr
                        pass
        # we loop around all columns so that we do this
        # column level work only once for each column
        for offsets in row_col_dict:
            col_letter = OpenPyxlCompositor._get_column_letter(offsets[1] + 1)
            ws.column_dimensions[col_letter].width = column_width
        # print("before callback function")
        if post_process_ws_func:
            post_process_ws_func(ws)

    @classmethod
    def to_xlsx_worksheet(
        cls,
        *,
        layout,
        worksheet,
        orientation="vertical",
        column_width=_DEFAULT_COLUMN_WIDTH,
        h_shift_by=1,
        v_shift_by=1,
        post_process_worksheet_func=None,
    ):

        """
        take a layout which contains a list of presentation models builts using the build_presentation_model function. this method is useful to control where the file is created and to add more attributes to the worksheet before it is being saved. updates the ws argument in place.

        args:
            layout: an nested list of presentation_models, examples: [presentation_model] or [presentation_model1, presentation_mode2] etc
            ws: openpyxl worksheet is which the presentation model will be rendered.
            orientation: if vertical, the top level presentation model elements are rendered vertically, and for every nested level the orientation is flipped.
                         if horizontal, then the behavior is inverse
            column-width: default=20, the default column width of all columns in the worksheet. individual column width cannot be set currently
            h_shift_by: defaulf=1, the no of horizontal rows that will be used while laying out the presentation model horizontally
            v_shift_by: defaulf=1, the no of vertical rows that will be used while laying out the presentation model vertically
            post_process_ws_func: a function that will be called back with the worksheet, for final processing. for example, if special formatting needs to be performed at the column level (freezing columns, hiding columns. etc.)
        """

        row_col_dict = cls._build_row_col_dict(
            layout,
            orientation=orientation,
            h_shift_by=h_shift_by,
            v_shift_by=v_shift_by,
        )
        cls._to_xlsx_worksheet(
            row_col_dict, worksheet, column_width, post_process_worksheet_func
        )

    @classmethod
    def to_xlsx(
        cls,
        *,
        layout,
        output_fp,
        orientation="vertical",
        column_width=_DEFAULT_COLUMN_WIDTH,
        h_shift_by=1,
        v_shift_by=1,
    ):
        """
        uses a layout which contains a list of presentation models built using the build_presentation_model function.

        args:
            layout: an nested list of presentation_models, examples: [presentation_model] or [presentation_model1, presentation_mode2] etc
            output_fp: the xlsx file name
            orientation: if vertical, the top level presentation model elements are rendered vertically, and for every nested level the orientation is flipped.
                         if horizontal, then the behavior is inverse
            column-width: default=20, the default column width of all columns in the worksheet. individual column width cannot be set currently
            h_shift_by: applied when `layout` has multiple presentation models. the value (default 1) is used to space the presentation models that are horizontal to each other
            v_shift_by: applied when `layout` has multiple presentation models. the value (default 1) is used to space the presentation models that are vertical to each other
        """

        row_col_dict = cls._build_row_col_dict(
            layout, orientation, h_shift_by, v_shift_by
        )
        workbook = Workbook()
        worksheet = workbook.active
        cls._to_xlsx_worksheet(
            row_col_dict,
            worksheet,
            column_width=column_width,
            post_process_ws_func=None,
        )

        workbook.save(output_fp)


class XlsxWriterCompositor(_XLSXCompositor):
    @staticmethod
    def _to_xlsx_worksheet(row_col_dict, ws, wb, column_width, post_process_ws_func):
        """
        Args:
            ws: worksheet to use for rendering
        """
        for offsets, (value, style, _) in row_col_dict.items():
            offsets = tuple(i + 1 for i in offsets)  # bump needed for openpyxl

            cell_format = wb.add_format(style.user_style)
            # the -1 is needed since XlsxWriter uses zero-based indexing
            if offsets[0] != offsets[2] or offsets[1] != offsets[3]:
                ws.merge_range(
                    first_row=offsets[0] - 1,
                    first_col=offsets[1] - 1,
                    last_row=offsets[2] - 1,
                    last_col=offsets[3] - 1,
                    data=df_type_to_str(value),
                    cell_format=cell_format,
                )
            else:
                ws.write(
                    offsets[0] - 1, offsets[1] - 1, df_type_to_str(value), cell_format
                )

        # we loop around all columns so that we do this
        # column level work only once for each column
        for offsets in row_col_dict:
            col_letter = _XLSXCompositor._get_column_letter(offsets[1] + 1)
            ws.set_column(col_letter + ":" + col_letter, column_width)
        if post_process_ws_func:
            post_process_ws_func(ws)

    @classmethod
    def to_xlsx_worksheet(
        cls,
        *,
        workbook,
        layout,
        worksheet,
        orientation="vertical",
        column_width=_DEFAULT_COLUMN_WIDTH,
        h_shift_by=1,
        v_shift_by=1,
        post_process_worksheet_func=None,
    ):

        """
        Take a layout which contains a list of presentation models builts using the build_presentation_model function. This method is useful to control where the file is created and to add more attributes to the worksheet before it is being saved. Updates the ws argument in place.

        Args:
            workbook: Workbook object needed for XlsxWriter to create format objects. Note that this parameter is not required for the equivalent OpenPyxlCompositor.
            layout: An nested list of presentation_models, examples: [presentation_model] or [presentation_model1, presentation_mode2] etc
            worksheet: openpyxl Worksheet is which the presentation model will be rendered.
            orientation: if vertical, the top level presentation model elements are rendered vertically, and for every nested level the orientation is flipped.
                         if horizontal, then the behavior is inverse
            column-width: default=20, the default column width of all columns in the worksheet. Individual column width cannot be set currently
            h_shift_by: defaulf=1, the no of horizontal rows that will be used while laying out the presentation model horizontally
            v_shift_by: defaulf=1, the no of vertical rows that will be used while laying out the presentation model vertically
            post_process_ws_func: a function that will be called back with the worksheet, for final processing. For example, if special formatting needs to be performed at the column level (freezing columns, hiding columns. etc.)
            kwargs: for future to options. currently not used
        """
        row_col_dict = cls._build_row_col_dict(
            layout,
            orientation=orientation,
            h_shift_by=h_shift_by,
            v_shift_by=v_shift_by,
        )
        cls._to_xlsx_worksheet(
            row_col_dict, worksheet, workbook, column_width, post_process_worksheet_func
        )

    @classmethod
    def to_xlsx(
        cls,
        *,
        layout,
        output_fp,
        orientation="vertical",
        column_width=_DEFAULT_COLUMN_WIDTH,
        h_shift_by=1,
        v_shift_by=1,
    ):
        """
        Uses a layout which contains a list of presentation models built using the build_presentation_model function.

        Args:
            layout: An nested list of presentation_models, examples: [presentation_model] or [presentation_model1, presentation_mode2] etc
            output_fp: the xlsx file name
            orientation: if vertical, the top level presentation model elements are rendered vertically, and for every nested level the orientation is flipped.
                         if horizontal, then the behavior is inverse
            column-width: default=20, the default column width of all columns in the worksheet. Individual column width cannot be set currently
             h_shift_by: applied when `layout` has multiple presentation models. The value (default 1) is used to space the presentation models that are horizontal to each other
             v_shift_by: applied when `layout` has multiple presentation models. The value (default 1) is used to space the presentation models that are vertical to each other


        The xlxswriter library seems to have better performance than the OpenPyxl library in some uses that were tested. For more information, run the benchmark/benchmarks.py provided with this library. Based on the desired performance and features needed the `engine` argument can be set accordingly.
        """

        row_col_dict = cls._build_row_col_dict(
            layout, orientation, h_shift_by, v_shift_by
        )
        workbook = xlsxwriter.Workbook(output_fp)
        worksheet = workbook.add_worksheet()
        cls._to_xlsx_worksheet(
            row_col_dict,
            worksheet,
            wb=workbook,
            column_width=column_width,
            post_process_ws_func=None,
        )
        workbook.close()


# For backward compatibility
class XLSXWriter(OpenPyxlCompositor):
    @staticmethod
    def new_workbook():
        """
        Helper function to create an empty workbook.
        """
        wb = Workbook()
        wb.remove(wb.active)
        return wb

    @staticmethod
    def add_sheet(wb: Workbook, sheet_name: str):
        """
        Given a wb and sheet_name, create a new sheet in the workbook
        and make this sheet the active sheet.
        """
        if len(sheet_name) > 31:  # 32+ chars will error when opened in excel
            raise RuntimeError(
                f"Sheet name must be 31 or fewer characters. {sheet_name=} is {len(sheet_name)} characters."
            )
        ws = wb.create_sheet(title=sheet_name)
        return ws

    @classmethod
    def to_xlsx_worksheet(
        cls,
        layout,
        ws,
        orientation="vertical",
        column_width=_DEFAULT_COLUMN_WIDTH,
        h_shift_by=1,
        v_shift_by=1,
        post_process_ws_func=None,
        **_,
    ):

        """
        Function that is provided for backward compatibility. New clients should call to_xlsx_worksheet on either OpenPyxlCompositor or XlsxWriterCompositor.
        """
        warnings.warn(
            "Consider switching to OpenPyxlCOmpositor.to_xlsx_worksheet",
            DeprecationWarning,
        )

        super().to_xlsx_worksheet(
            layout=layout,
            worksheet=ws,
            orientation=orientation,
            column_width=column_width,
            h_shift_by=h_shift_by,
            v_shift_by=v_shift_by,
            post_process_worksheet_func=post_process_ws_func,
        )

    @classmethod
    def to_xlsx(
        cls,
        layout,
        output_fp,
        orientation="vertical",
        column_width=_DEFAULT_COLUMN_WIDTH,
        h_shift_by=1,
        v_shift_by=1,
        **_,
    ):

        """
        Function that is provided for backward compatibility. New clients should call to_xlsx on either OpenPyxlCompositor or XlsxWriterCompositor.
        """

        warnings.warn(
            "Consider switching to OpenPyxlCOmpositor.to_xlsx_worksheet",
            DeprecationWarning,
        )

        super().to_xlsx(
            layout=layout,
            output_fp=output_fp,
            orientation=orientation,
            column_width=column_width,
            h_shift_by=h_shift_by,
            v_shift_by=v_shift_by,
        )

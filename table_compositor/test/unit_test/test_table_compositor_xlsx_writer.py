import os
import sys
import tempfile
import unittest
from functools import partial
from collections import namedtuple

from openpyxl import load_workbook

import table_compositor.test.unit_test.table_compositor_fixtures as tcfx

TestFixtures = namedtuple('TestFixtures',
        ['name', 'fixture_func', 'grid', 'nested'])

def get_expected_output_folder(fname):
    base_path = os.path.join(
            os.path.dirname(__file__),
            '..', 'expected')
    os.makedirs(base_path, exist_ok=True)
    expected_fp = os.path.join(base_path, fname)
    return expected_fp

class TestXlsxWriterMeta(type):
    def __new__(cls, names, bases, attrs):
        def test_func_constructor(name, func_to_call, engine, orientation):
            def _func(self):
                layout = func_to_call()
                # we drop the engine name from the test, since the expected file is the same for both engines
                fname = name.replace('_' + engine.__name__, '') + ".xlsx"
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_dir = tempfile.mkdtemp(dir=tempfile.gettempdir())
                    output_fp = os.path.join(
                            temp_dir, fname)
                    engine.to_xlsx(
                            layout=layout,
                            output_fp=output_fp,
                            orientation=orientation)
                    expected_fp = get_expected_output_folder(fname)
                    #use this to create the initial expected output
                    # import shutil
                    # shutil.copyfile(output_fp, expected_fp)
                    self.compare(expected_fp, output_fp)

            return _func

        scenarios = tcfx.get_scenarios()

        for s in scenarios:
            attrs[s.name] = test_func_constructor(s.name, partial(
                    s.fixture_func,
                    grid=s.grid,
                    nested=s.nested,
                    callback_func_cls=s.engine_and_callback_funcs_cls[1]),
                    s.engine_and_callback_funcs_cls[0],
                    s.orientation)

        return type.__new__(cls, names, bases, attrs)

class TestUnit(unittest.TestCase , metaclass=TestXlsxWriterMeta):
    '''
    All tests will be produced by the meta class
    '''
    def compare(self, expected_fp, output_fp):

        ews = load_workbook(expected_fp).active
        ows = load_workbook(output_fp).active

        for i in range(1, ews.max_row):
            for j in range(1, ews.max_column):

                e_cell = ews.cell(row=i, column=j)
                o_cell = ows.cell(row=i, column=j)

                self.assertEqual(
                        e_cell.value, o_cell.value)

                self.assertEqual(
                        e_cell.number_format,
                        o_cell.number_format)

                # this test not compatible between openpyxl and xlsxwriter
                # self.assertEqual(
                #     e_cell.fill.fgColor.rgb,
                #     o_cell.fill.fgColor.rgb)

                self.assertEqual(
                        e_cell.border.left.style,
                        o_cell.border.left.style)


if __name__ == '__main__':
    unittest.main()
    # t = TestUnit()

    # t.test_get_multi_hierarchical_df_with_layouts_grid_True_nested_True_orientation_vertical()
    # t.test_get_multi_hierarchical_df_with_layouts_grid_True_nested_True_orientation_horizontal

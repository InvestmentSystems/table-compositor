import os
import tempfile
import unittest
from functools import partial

from openpyxl import load_workbook

from table_compositor.test.unit_test import fixtures


def get_expected_output_folder(fname):
    base_path = os.path.join(os.path.dirname(__file__), "..", "expected")
    os.makedirs(base_path, exist_ok=True)
    expected_fp = os.path.join(base_path, fname)
    return expected_fp


class TestUnit(unittest.TestCase):
    def _compare(self, expected_fp, output_fp):

        ews = load_workbook(expected_fp).active
        ows = load_workbook(output_fp).active

        for i in range(1, ews.max_row):
            for j in range(1, ews.max_column):

                e_cell = ews.cell(row=i, column=j)
                o_cell = ows.cell(row=i, column=j)

                self.assertEqual(e_cell.value, o_cell.value)

                self.assertEqual(e_cell.number_format, o_cell.number_format)

                # this test not compatible between openpyxl and xlsxwriter
                # self.assertEqual(
                #     e_cell.fill.fgColor.rgb,
                #     o_cell.fill.fgColor.rgb)

                self.assertEqual(e_cell.border.left.style, o_cell.border.left.style)

    def _test_helper(self, func_to_call, name, engine, orientation):
        layout = func_to_call()
        # we drop the engine name from the test, since the expected file is the same for both engines
        fname = name.replace("_" + engine.__name__, "") + ".xlsx"

        # this is a quick hack to have the same expected fp for for pandas and static_frame
        expected_fname = fname.replace("_static_frame", "").replace("_pandas", "")
        with tempfile.TemporaryDirectory(suffix="table_compositor") as root_temp_dir:
            temp_dir = tempfile.mkdtemp(dir=root_temp_dir)
            output_fp = os.path.join(temp_dir, fname)
            engine.to_xlsx(layout=layout, output_fp=output_fp, orientation=orientation)
            expected_fp = get_expected_output_folder(expected_fname)
            # use this to create the initial expected output
            # shutil.copyfile(output_fp, expected_fp)
            self._compare(expected_fp, output_fp)

    def test_xlsx_writer(self):
        scenarios = fixtures.get_scenarios()

        for s in scenarios:
            func_to_call = partial(
                s.fixture_func,
                grid=s.grid,
                nested=s.nested,
                callback_func_cls=s.engine_and_callback_funcs_cls[1],
                frame_library=s.frame_library,
            )

            self._test_helper(
                func_to_call,
                s.name,
                s.engine_and_callback_funcs_cls[0],
                s.orientation,
            )


if __name__ == "__main__":
    unittest.main()
